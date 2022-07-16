import win32com.client as win32
import pandas as pd
import openpyxl
import pdb
def CrawlOutlook(email):
    if(email.Unread ==True):
        if (email.SenderEmailAddress =='aee-gs-order-spike-report@amazon.com'):
            email.Unread = False
            file = open(download_path, mode='w')
            file.writelines(email.Body)
            file.close()
            return 1
    else:
        return 0
def JudgeExcel(checker,filepath):
    if(checker ==1):
        excel = pd.read_csv(download_path)
        # print(excel.all(axis=None))
        # print(excel)
        return excel
    else:
        excel = pd.DataFrame({'ASIN': [0], 'Destination': [0], 'Source': [0]})
        # print(excel.all(axis=None))
        return excel
def Handler(container):
    j=0
    limit=[]
    asin = []
    destination = []
    source = []
    inventory =[]
    units =[]
    deal_end_time=[]
    # print(container.iloc[0,0])
    if(container.all(axis=None)):
        for i in range(len(container)):
            if (container.iloc[i, 0][0:4] == 'ASIN' and j < 2):
                limit.append(container.index[i])
                # print(limit[j])
                j += 1
        for i in range(limit[j - 1]):
            if (i > 0):
                k = 0
                tab_collector = []
                index_k = 0
                # print(container.iloc[i, 0])
                for cha in container.iloc[i, 0]:
                    if (cha == '\t'):
                        index_k += 1
                        # print(k)
                        tab_collector.append(k)
                        if (index_k == 1):
                            asin.append(container.iloc[i, 0][0:k])
                            # print(container.iloc[i, 0][0:k])
                            k_position1 = k
                        if (index_k == 2):
                            destination.append(container.iloc[i, 0][k_position1 + 1:k])
                            k_position2 = k
                        if (index_k == 3):
                            source.append(container.iloc[i, 0][k_position2 + 1:k])
                            k_position3 = k
                        if (index_k ==4):
                            inventory.append(container.iloc[i,0][k_position3+1:k])
                            k_position4 = k
                        if (index_k == 5):
                            units.append(container.iloc[i,0][k_position4+1:k])
                            k_position5 = k
                        if (index_k == 6):
                            deal_end_time.append(container.iloc[i,0][k_position5+1:k])
                    k += 1
        data = pd.DataFrame({'ASIN': asin, 'Destination': destination, 'Source': source,'Inventory':inventory,'Units':units,'Deal End Time':deal_end_time})
        # print(data)
        return data
    else:
        data = pd.DataFrame({'ASIN': [0], 'Destination': [0], 'Source': [0],'Inventory': [0],'Units':[0],'Deal End Time':[0]})
        return data
def AddToExcel(data):
    if(data.all(axis=None)):
        data = Filter(data,destination,source)
        Excel_name = r'C:\Users\xinyyang\Desktop\UK-US_Selection_Health\UNYANK\Pre.xlsx'
        open_excel = openpyxl.load_workbook(Excel_name)
        sheet =open_excel['Pretreatment']
        # print(data)
        initial_index = len(tuple(sheet.rows))
        # print(initial_index)
        for i in range(len(data.index)):
            sheet.cell(i + 1 + initial_index, 1, data.iloc[i][0])
            sheet.cell(i + 1 + initial_index, 2, data.iloc[i][1])
            sheet.cell(i + 1 + initial_index, 3, data.iloc[i][2])
            sheet.cell(i + 1 + initial_index, 4, data.iloc[i][3])
            sheet.cell(i + 1 + initial_index, 5, data.iloc[i][4])
            sheet.cell(i + 1 + initial_index, 6, data.iloc[i][5])
        open_excel.save(Excel_name)
        open_excel.close()
def DeleteExcel(excel_name,sheet_name):
    open_excel = openpyxl.load_workbook(excel_name)
    sheet = open_excel[sheet_name]
    initial_index = len(tuple(sheet.rows))
    # print(initial_index)
    for i in range(initial_index-1):
        sheet.cell(i+2,1).value = None
        sheet.cell(i+2,2).value = None
        sheet.cell(i+2,3).value = None
        sheet.cell(i+2,4).value = None
        sheet.cell(i+2,5).value = None
        sheet.cell(i+2,6).value = None
    open_excel.save(excel_name)
    open_excel.close()
def Filter(data,destinationscope,sourcescope):
    list=[]
    # print(data)\
    # print(type(data.iloc[2]['Destination']),data.iloc[2]['Destination'])
    for i in range(len(data.index)):
        if((data.iloc[i]['Destination'] not in destinationscope) or (data.iloc[i]['Source'] not in sourcescope )):
            list.append(i)
    # print(list)
    data.drop(list,inplace=True)
    data.reset_index(inplace=True,drop=True)
    # print(data)
    return data

source = ['104444012','771770','111172','338801','338811','44571','3240','1','3','4','6']
destination = ['104444012','338801','338811','111172','771770','3240','1','3']
outlook = win32.Dispatch('outlook.application').GetNamespace("MAPI")
Inbox = outlook.GetDefaultFolder(6)
Unyank1 = Inbox.Folders['Unyank1']
messages = Unyank1.Items
content = Unyank1.Items[1].Body
#download_path没有拷到新class中，想少用一个excel
download_path = r'C:\Users\xinyyang\Desktop\UK-US_Selection_Health\UNYANK\outlook\2.csv'
Excel_name = r'C:\Users\xinyyang\Desktop\UK-US_Selection_Health\UNYANK\Pre.xlsx'
#清空上次残留数据
DeleteExcel(Excel_name,'Pretreatment')
headers = ['ASIN','Destination','Source','Inventory','Units','Deal End Time']
df = pd.DataFrame(columns=headers)
for message in messages:
    checker = CrawlOutlook(message)
    pretreatment = JudgeExcel(checker,download_path)
    # print(pretreatment)
    datainfo = Handler(pretreatment)
    # print(datainfo)
    if(datainfo.all(axis=None)):
        df=df.append(datainfo)

#又把excel打开一遍，读取信息至dataframe中，然后去重，存到excel中（更简单的办法没有成功，先实现功能再说）
# raw_data = pd.read_excel(Excel_name,sheet_name='Pretreatment')
# raw_data.drop_duplicates(subset=['ASIN','Destination','Source'])
# DeleteExcel(Excel_name,'Pretreatment')
# print(raw_data)
# AddToExcel(raw_data)
# df.to_excel('test.xlsx',index=False)
# pdb.set_trace()
df = df.reset_index(drop=True)
df = df.drop_duplicates(subset=['ASIN','Destination','Source'])
AddToExcel(df)
# print(df)