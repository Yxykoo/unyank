import pandas as pd
import openpyxl
import os
import time
import shutil
from datetime import date
#这个list的数据要改成从AgsOperation那个函数传回来
list_test = [['US', 'DE', 'https://123.amazon.com/123=1']]
id_list_test = []
class Record:
    def __init__(self,request_id_list):
        self.request_id_list = request_id_list
        self.address = r'C:\Users\xinyyang\Desktop\UK-US_Selection_Health\UNYANK\New folder\Pre.xlsx'
        self.backup_file = r'C:\Users\xinyyang\Desktop\UK-US_Selection_Health\UNYANK\New folder\backup'
        self.sheet_name1 = 'Result'
        self.sheet_name2 = 'work_history'
        self.subtask1 = 'GSgatewaypromotion'
        self.subtask2 = 'AgOrderSpikeAndKillScripts'
    def Sort(self,dataframe):
        id_list = []
        arc_list = []
        subtask_list = []
        today_list = []
        for i in range(len(dataframe.index)):
            for j in range(len(self.request_id_list)):
                if(dataframe.iloc[i]['Source']==self.request_id_list[j][0] and dataframe.iloc[i]['Destination']==self.request_id_list[j][1]):
                    id_list.append(self.request_id_list[j][2])
                    continue
            # import pdb
            # pdb.set_trace()
            arc_list.append(dataframe.iloc[i]['Source']+'-'+dataframe.iloc[i]['Destination'])
            subtask_list.append(self.subtask1)
            today_list.append(date.today().strftime("%m/%d/%Y"))
        print(dataframe)
        print(id_list)
        dataframe.insert(3,'Request_id',id_list,allow_duplicates=True)
        dataframe.insert(3,'ARC',arc_list,allow_duplicates=True)
        dataframe.insert(4,'Subtask',subtask_list,allow_duplicates=True)
        dataframe.insert(0,'Date',today_list,allow_duplicates=True)
        return dataframe

    def AddToExcel(self,dataframe):
        open_excel = openpyxl.load_workbook(self.address)
        sheet = open_excel[self.sheet_name2]
        # print(sheet.max_row)
        start_index = sheet.max_row +1
        for i in range(len(dataframe.index)):
            sheet.cell(start_index+i,1,dataframe.iloc[i][0])
            sheet.cell(start_index+i,2,dataframe.iloc[i][1])
            sheet.cell(start_index + i, 3, dataframe.iloc[i][2])
            sheet.cell(start_index + i, 4, dataframe.iloc[i][3])
            sheet.cell(start_index + i, 5, dataframe.iloc[i][4])
            sheet.cell(start_index + i, 6, dataframe.iloc[i][5])
            sheet.cell(start_index + i, 8, dataframe.iloc[i][6])
        open_excel.save(self.address)
        open_excel.close()

    def Conversion(self,data):
        for i in range(len(data.index)):
            if(data.iloc[i]['Destination'] =='KSA'):
                data.at[i,'Destination']= 'SA'
        return data
    def Operation(self):
        shutil.copy(self.address, self.backup_file)
        data = pd.read_excel(self.address,sheet_name=self.sheet_name1)
        #在进行sort前要把dataframe中的KSA改成SA
        data = self.Conversion(data)
        data = self.Sort(data)
        # print(data)
        self.AddToExcel(data)
        print('Record request_id Done')
if __name__ == '__main__':
    # record = Record(list_test)
    # record.Operation()
    print(len(id_list_test))
    print(list(set(id_list_test)))
# print()