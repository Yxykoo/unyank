import pandas as pd
import openpyxl
import sys
class ArcChanger:
    def __init__(self):
        self.destination_scope = ['SG', 'AE', 'KSA', 'AU', 'MX', 'CN', 'US', 'UK','DE','TR']
        self.destination_id = [104444012,338801,338811,111172,771770,3240,1,3,4,338851]
        self.source_scope = ['SG', 'MX', 'AU', 'AE', 'KSA', 'IN', 'CN', 'US', 'UK', 'DE', 'JP']
        self.source_id = [104444012,771770,111172,338801,338811,44571,3240,1,3,4,6]
        self.excel_name = r'C:\Users\xinyyang\Desktop\UK-US_Selection_Health\UNYANK\New folder\Pre.xlsx'
        self.sheet_name1 = 'Pretreatment'
        self.sheet_name2 = 'Result'
    def replace_source(self,name):
        name['Source'] = name['Source'].replace(self.source_id, self.source_scope)
        return name

    def replace_destination(self,name):
        name['Destination'] = name['Destination'].replace(self.destination_id, self.destination_scope)
        return name

    def DeleteExcel(self):
        open_excel = openpyxl.load_workbook(self.excel_name)
        sheet = open_excel[self.sheet_name2]
        initial_index = len(tuple(sheet.rows))
        # print(initial_index)
        for i in range(initial_index-1):
            sheet.cell(i+2,1).value = None
            sheet.cell(i+2,2).value = None
            sheet.cell(i+2,3).value = None
        open_excel.save(self.excel_name)
        open_excel.close()
    def Filter(self,data):
        list=[]
        # print(data)
        for i in range(len(data.index)):
            if((data.iloc[i]['Destination'] not in self.destination_scope) or (data.iloc[i]['Source'] not in self.source_scope )):
                list.append(i)
        # print(list)
        data.drop(list,inplace=True)
        data.reset_index(inplace=True,drop=True)
        # print(data)
        return data
    def AddExcel(self):
        file = pd.read_excel(self.excel_name,self.sheet_name1)
        file = self.replace_source(file)
        file = self.replace_destination(file)
        # file = self.Filter(file)
        if file.shape ==0:
            print('no ASINs available')
            sys.exit()
        self.DeleteExcel()
        open_excel = openpyxl.load_workbook(self.excel_name)
        sheet = open_excel[self.sheet_name2]
        # import pdb
        # pdb.set_trace()
        for i in range(len(file.index)):
            sheet.cell(i + 2, 1, file.iloc[i][0])
            sheet.cell(i + 2, 2, file.iloc[i][2])
            sheet.cell(i + 2, 3, file.iloc[i][1])
        open_excel.save(self.excel_name)
        open_excel.close()
        print('Save Email info into excel Done')
if __name__ == '__main__':
    arc_changer = ArcChanger()
    arc_changer.AddExcel()